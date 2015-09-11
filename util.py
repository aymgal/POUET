"""
Useful functions and definitions
"""
import astropy.coordinates.angles as angles
from astropy.table import Table
from astropy.time import Time
from astropy import units as u
import urllib2
import re
import getopt, sys
import openpyxl  # see http://openpyxl.readthedocs.org/en/latest/index.html
import obs
from bisect import bisect_left

import csv
import numpy as np

# La Silla Telescope Parameters
def get_telescope_params():
	# LaSilla
	lat=angles.Angle("-29d15m33.7s")
	lon=angles.Angle("-70.7313d")
	elev = 2400

	return lat, lon, elev

def get_AzAlt(alpha, delta, obs_time=Time.now(), ref_dir=0):

	"""
	idea from http://aa.usno.navy.mil/faq/docs/Alt_Az.php

	Compute the azimuth and altitude of a source at a given time (by default current time of 
	execution), given its alpha and delta coordinates.

	WARNING ! Azimuth and altitude are computed at La Silla Observatory

	"""

	lat, lon, elev = get_telescope_params()

	# Untouched code from Azimuth.py
	D = obs_time.jd - 2451545.0
	GMST = 18.697374558 + 24.06570982441908*D
	epsilon= np.deg2rad(23.4393 - 0.0000004*D)
	eqeq= -0.000319*np.sin(np.deg2rad(125.04 - 0.052954*D)) - 0.000024*np.sin(2.*np.deg2rad(280.47 + 0.98565*D))*np.cos(epsilon)
	GAST = GMST + eqeq
	GAST -= np.floor(GAST/24.)*24.

	LHA = angles.Angle((GAST-alpha.hour)*15+lon.degree, unit="degree")
	if LHA > 0: LHA += angles.Angle(np.floor(LHA/360.)*360., unit="degree")
	else: LHA -= angles.Angle(floor(LHA/360.)*360., unit="degree")

	sina=np.cos(LHA.radian)*np.cos(delta.radian)*np.cos(lat.radian)+np.sin(delta.radian)*np.sin(lat.radian)
	Alt = angles.Angle(np.arcsin(sina),unit="radian")

	num = -np.sin(LHA.radian)
	den = np.tan(delta.radian)*np.cos(lat.radian)-np.sin(lat.radian)*np.cos(LHA.radian)

	Az = angles.Angle(np.arctan2(num,den), unit="radian")
	Az-=angles.Angle(ref_dir, unit="degree")

	# I changed this to get the same angle as the edp, using 0 (North) as reference
	if Az.degree < 0:
		Az+=angles.Angle(360, unit="degree")

	return Az, Alt



def reformat(coordinate, format):
	"""
	Transform a coordinate (hour, degree) in the format of your choice

	HHhDDdSSs <---> HH:DD:SS

	update : apparently useless, can use angle objects instead
	"""

	if 'm' in coordinate:
		if 'd' in coordinate:
			hd = coordinate.split('d')[0]
			m = coordinate.split('d')[1].split('m')[0]
		if 'h' in coordinate:
			hd = coordinate.split('h')[0]
			m = coordinate.split('h')[1].split('m')[0]
		s  = coordinate.split('m')[1].split('s')[0]

	elif ':' in coordinate:
		[hd, m, s] = coordinate.split(':')

	else:
		raise ValueError("%s, Unknown coordinate input format!" %coordinate)

	if format == 'numeric':
		return "%s:%s:%s" % (hd, m, s)

	elif format == 'alphabetic_degree':
		return "%sd%sm%ss" % (hd, m, s)

	elif format == 'alphabetic_hour':
		return "%sh%sm%ss" % (hd, m, s)

	else:
		raise ValueError("%s, Unknown coordinate output format!" %format)


def takeclosest(dict, key, value):
	"""
	Assumes dict[key] is sorted. Returns the dict value which dict[key] is closest to value.
	If two dict[key] are equally close to value, return the highest (i.e. latest).
	This is much faster than a simple min loop, although a bit more tedious to use.
	"""
	mylist = [elt[key] for elt in dict]

	pos = bisect_left(mylist, value)
	if pos == 0:
		return dict[0]
	if pos == len(mylist):
		return dict[-1]
	before = dict[pos - 1]
	after = dict[pos]
	if after[key] - value <= value - before[key]:
		return after
	else:
		return before


def hilite(string, status, bold):
	'''Graphism: colors and bold in the terminal'''

	if not sys.stdout.isatty() : return '*'+string+'*'

	attr = []
	if status:
		# green
		attr.append('32')
	else:
		# red
		attr.append('31')
	if bold:
		attr.append('1')
	return '\x1b[%sm%s\x1b[0m' % (';'.join(attr), string)

def excelimport(filename, obsprogram=None):
	"""
	Import an excel catalog(...) into a list of observables
	I directly read the excel values, I do NOT evaluate the formulas in them.
	It is up to you to put the right mjd in the excel sheets.

	Warning : NEVER use the coordinates from an excel sheet to create an rdb night planning. ALWAYS use the rdb catalogs loaded in the edp. And double check the distance to moon !
	"""

	observables = []

	#### For BEBOP
	if obsprogram == 'bebop':
		"""
		special properties:

		phases : a list of dictionnaries : [{mjd, phase, hourafterstart }]
		comment : a string of comments (exptime, requested phase,...)
		internalobs : a boolean (0 or 1), allowing or not observability
		"""

		try:
			wb = openpyxl.load_workbook(filename, data_only=True)  # Read the excel spreadsheet, loading the values directly
			#ws = wb.active  # choose active sheet
			ws = wb['Sheet1']
		except:
			raise RuntimeError("Either %s does not exists, or it is not in .xlsx format !!" % filename)

		# Get tabler limits
		rows = ws.rows
		columns = ws.columns
		breakcolind = None
		breakrowind = None
		for ind, cell in enumerate(rows[1]):
			if cell.value == None:
				#breakcolind = cell.column
				breakcolind = rows[1][ind-1].column
				break
			else:
				pass

		for ind, cell in enumerate(columns[0]):
			if cell.value == None:
				#breakrowind = cell.row
				breakrowind = columns[0][ind-1].row
				break
			else:
				pass

		# Read only the non "None" data and put it in a table of dict, because fuck excel and fuck openpyxl.
		data = ws['A1':'%s%s' % (breakcolind, breakrowind)]
		"""
		Structure of the spreadsheet:
		Infos are from A1 to W2
		Datas are from A3 ro W30
		B1 : actual modified julian date
		A : name
		B : target
		C : comment
		I : observability
		J : requested phase
		M1 to W1 : mjd over the night
		M2 to W2 : corresponding time after night start, in hours
		M to W : phases
		"""

		phasesnames = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
		values = {}
		for indr, row in enumerate(data):
			for indc, cell in enumerate(row):
				try:
					values[cell.address] = cell.value # Apparently, this is an old version
				except:
					values[cell.coordinate] = cell.value

		for i in np.arange(3, 31):
			# create an observable object with the common properties
			name = values['A%s' % str(i)]
			minangletomoon = 70
			maxairmass = 1.5
			exptime = 1800  # THIS IS NOT ALWAYS THE CASE

			coordinates = values['B%s' % str(i)]
			alpha = coordinates[0:2]+':'+coordinates[2:4]+':'+coordinates[4:6]
			delta = coordinates[7:9]+':'+coordinates[9:11]+':'+coordinates[11:13]
			if coordinates[6] == "S":
				delta = '-'+delta

			observable = obs.Observable(name=name, obsprogram=obsprogram, alpha=alpha, delta=delta, minangletomoon=minangletomoon, maxairmass=maxairmass, exptime=exptime)

			# add properties specific to this program
			## Tricky stuff here : the jdb in the excel sheet is the mjd + 0.5.
			phases = [{'mjd': values['%c%i' % (col, 1)]-0.5, 'hourafterstart': values['%c%i' % (col, 2)], 'phase': values['%c%i' % (col, i)]} for col in phasesnames]

			observable.phases = phases
			if values['I%s' % str(i)] == 'yes':
				observable.internalobs = 1
			else:
				observable.internalobs = 0
			comment = ''
			if values['C%s' % str(i)] is not None:
				comment = comment + values['C%s' % str(i)]
			if values['J%s' % str(i)] != '/':
				comment = comment + '\n' + 'Requested phase: ' + values['J%s' % str(i)]
			if comment != '':
				observable.comment = comment
			observables.append(observable)

		#TODO: check that the modified julian date corresponds to the ongoing night
		#TODO: assert that the above structure is correct ! (use keywords in the Info fields...?)

	if obsprogram == "transit":
		pass

	if obsprogram == "superwasp":
		# http://openpyxl.readthedocs.org/en/latest/optimized.html --- that will be useful for Amaury's monstruous spreadsheet
		"""
		special properties:

		phases : a list of dictionnaries : [{mjd, phase, hourafterstart }]
		comment : a string of comments (exptime, requested phase,...)
		internalobs : a boolean (0 or 1), allowing or not observability
		"""

		print 'reading %s...' % filename
		try:
			wb = openpyxl.load_workbook(filename, data_only=True)  # Read the excel spreadsheet, loading the values directly
			ws = wb['Observations']  # choose active sheet
		except:
			raise RuntimeError("Either %s does not exists, or it is not in .xlsx format !!" % filename)

		print 'get tabler limits...'
		# Get tabler limits
		rows = ws.rows
		columns = ws.columns
		breakcolind = None
		breakrowind = None
		for ind, cell in enumerate(rows[1]):
			if cell.value == None:
				#breakcolind = cell.column
				breakcolind = rows[1][ind-1].column
				break
			else:
				pass

		for ind, cell in enumerate(columns[0]):
			if cell.value == None:
				#breakrowind = cell.row
				breakrowind = columns[0][ind-1].row
				break
			else:
				pass
		print breakrowind, breakcolind
		sys.exit()


		# Read only the non "None" data and put it in a table of dict, because fuck excel and fuck openpyxl.
		data = ws['A1':'%s%s' % (breakcolind, breakrowind)]
		"""
		Structure of the spreadsheet:
		Infos are from A1 to W2
		Datas are from A3 ro W30
		B1 : actual modified julian date
		A : name
		B : target
		C : comment
		I : observability
		J : requested phase
		M1 to W1 : mjd over the night
		M2 to W2 : corresponding time after night start, in hours
		M to W : phases
		"""

	if obsprogram == "followup":
		pass

	return observables

def rdbimport(filepath, obsprogram, col_name, col_alpha, col_delta, return_all=False):

	f = open(filepath, 'rb')
	reader = csv.reader(f, delimiter='\t')
	headers = reader.next()
	
	# print Table.read(filepath, format='rdb') # This will be removed in next version of astropy so
	# let's do it in another way:
	cat = np.recfromtxt(filepath, names=headers, comments='--', delimiter='\t')
	cat = Table(cat[1:], names=cat[0])

	observables = []
	for ii, li in enumerate(cat):
		name, alpha, delta = li[col_name], li[col_alpha], li[col_delta]
		obj = cat[ii]
		observables.append(obs.Observable(name=name, obj=obj, obsprogram=obsprogram, alpha=alpha, delta=delta))

	if return_all:
		return observables, cat
	else:
		return observables
